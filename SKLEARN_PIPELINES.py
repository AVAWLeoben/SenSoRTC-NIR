# -*- coding: utf-8 -*-
"""
SKLearn NIR pipelines for SenSoRTC.

Supports mixed training inputs:
    - EVK SQALAR .mat files
    - SenSoRTC .npy raw spectral chunks
    - compressed .npz spectral arrays
    - Excel .xlsx/.xlsm files containing valid spectra for one class

Canonical spectral cube layout:
    (width, bands, lines), e.g. (312, 220, n_lines)

Training format after loading:
    X = (n_spectra, bands)
    y = (n_spectra,)

Design:
    NIRPipelineBase contains reusable loading/training/evaluation/saving logic.
    Concrete subclasses only define build_pipeline().
"""

from pathlib import Path

import joblib
import numpy as np
from scipy.io import loadmat

from sklearn.pipeline import Pipeline
from sklearn.base import BaseEstimator, ClassifierMixin, TransformerMixin
from sklearn.preprocessing import StandardScaler
from sklearn.neural_network import MLPClassifier
from sklearn.svm import SVC, LinearSVC
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report, confusion_matrix
from sklearn.feature_selection import mutual_info_classif
from sklearn.decomposition import PCA

from scipy_transformer_adapter import SavGolTransformer


class MRMRBandSelector(BaseEstimator, TransformerMixin):
    """Greedy mRMR-style spectral band selector.

    Selects n_features bands with high mutual information to the class labels
    while penalising redundancy to already-selected bands.

    Score at each step:
        relevance(candidate) - redundancy_weight * mean_abs_corr(candidate, selected)

    The selected band indices are learned during fit() and then reused during
    inference, so live camera spectra are reduced to exactly the same bands.
    """

    def __init__(
        self,
        n_features=30,
        redundancy_weight=1.0,
        random_state=42,
        max_samples_for_fit=20000,
    ):
        self.n_features = int(n_features)
        self.redundancy_weight = float(redundancy_weight)
        self.random_state = random_state
        self.max_samples_for_fit = int(max_samples_for_fit)

    def fit(self, X, y):
        X = np.asarray(X, dtype=np.float32)
        y = np.asarray(y)

        if X.ndim != 2:
            raise ValueError(f"MRMRBandSelector expects 2D X, got shape={X.shape}")

        n_samples, n_bands = X.shape
        n_select = max(1, min(self.n_features, n_bands))

        if n_samples > self.max_samples_for_fit:
            rng = np.random.default_rng(self.random_state)
            idx = rng.choice(n_samples, self.max_samples_for_fit, replace=False)
            X_fit = X[idx]
            y_fit = y[idx]
        else:
            X_fit = X
            y_fit = y

        relevance = mutual_info_classif(
            X_fit,
            y_fit,
            discrete_features=False,
            random_state=self.random_state,
        )
        relevance = np.nan_to_num(relevance, nan=0.0, posinf=0.0, neginf=0.0).astype(np.float32)

        # Correlation is used only for redundancy between spectral bands.
        Xc = X_fit - np.mean(X_fit, axis=0, keepdims=True)
        std = np.std(Xc, axis=0, keepdims=True) + 1e-8
        Xz = Xc / std

        selected = []
        remaining = list(range(n_bands))

        first = int(np.argmax(relevance))
        selected.append(first)
        remaining.remove(first)

        while len(selected) < n_select and remaining:
            sel_arr = np.asarray(selected, dtype=int)
            rem_arr = np.asarray(remaining, dtype=int)

            # mean absolute correlation of every remaining band to selected bands
            corr = np.abs((Xz[:, rem_arr].T @ Xz[:, sel_arr]) / max(1, Xz.shape[0] - 1))
            redundancy = np.mean(corr, axis=1)

            score = relevance[rem_arr] - self.redundancy_weight * redundancy
            best_remaining_idx = int(np.argmax(score))
            best_band = int(rem_arr[best_remaining_idx])

            selected.append(best_band)
            remaining.remove(best_band)

        self.selected_bands_ = np.asarray(selected, dtype=np.int64)
        self.selected_bands_sorted_ = np.sort(self.selected_bands_)
        self.n_features_in_ = int(n_bands)
        self.relevance_ = relevance

        print(f"[MRMR] selected {len(self.selected_bands_)} / {n_bands} bands:")
        print(self.selected_bands_.tolist())

        return self

    def transform(self, X):
        X = np.asarray(X, dtype=np.float32)

        if not hasattr(self, "selected_bands_"):
            raise RuntimeError("MRMRBandSelector is not fitted yet.")

        if X.ndim != 2:
            raise ValueError(f"MRMRBandSelector expects 2D X, got shape={X.shape}")

        if X.shape[1] != self.n_features_in_:
            raise ValueError(
                f"Band-count mismatch in MRMRBandSelector: got {X.shape[1]}, "
                f"expected {self.n_features_in_}."
            )

        return X[:, self.selected_bands_]

    def get_selected_bands(self):
        if not hasattr(self, "selected_bands_"):
            return None
        return self.selected_bands_.copy()


class SNVTransformer(BaseEstimator, TransformerMixin):
    """Standard Normal Variate transformer for spectra.

    Unlike dataset z-score/StandardScaler, SNV works per spectrum:
        X_snv[i] = (X[i] - mean(X[i])) / std(X[i])

    This helps reduce spectrum-wide brightness/scatter variation while keeping
    the live inference path identical to training because the transformer is
    stored inside the sklearn Pipeline.
    """

    def __init__(self, eps=1e-8):
        self.eps = float(eps)

    def fit(self, X, y=None):
        X = np.asarray(X, dtype=np.float32)
        if X.ndim != 2:
            raise ValueError(f"SNVTransformer expects 2D X, got shape={X.shape}")
        self.n_features_in_ = int(X.shape[1])
        return self

    def transform(self, X):
        X = np.asarray(X, dtype=np.float32)
        if X.ndim != 2:
            raise ValueError(f"SNVTransformer expects 2D X, got shape={X.shape}")
        if hasattr(self, "n_features_in_") and X.shape[1] != self.n_features_in_:
            raise ValueError(
                f"Band-count mismatch in SNVTransformer: got {X.shape[1]}, "
                f"expected {self.n_features_in_}."
            )
        mean = np.mean(X, axis=1, keepdims=True)
        std = np.std(X, axis=1, keepdims=True)
        return (X - mean) / (std + self.eps)


class PCALoadingBandSelector(BaseEstimator, TransformerMixin):
    """Select original spectral bands by PCA loading strength.

    This is a feature selector, not a PCA projection. It fits PCA on the
    spectra, scores each original band by the absolute loadings of the first
    principal components, and then keeps the strongest original bands.
    """

    def __init__(
        self,
        n_features=30,
        n_components=5,
        random_state=42,
        max_samples_for_fit=20000,
        weight_by_variance=True,
    ):
        self.n_features = int(n_features)
        self.n_components = int(n_components)
        self.random_state = random_state
        self.max_samples_for_fit = int(max_samples_for_fit)
        self.weight_by_variance = bool(weight_by_variance)

    def fit(self, X, y=None):
        X = np.asarray(X, dtype=np.float32)
        if X.ndim != 2:
            raise ValueError(f"PCALoadingBandSelector expects 2D X, got shape={X.shape}")

        n_samples, n_bands = X.shape
        n_select = max(1, min(self.n_features, n_bands))
        n_components = max(1, min(self.n_components, n_samples, n_bands))

        if n_samples > self.max_samples_for_fit:
            rng = np.random.default_rng(self.random_state)
            idx = rng.choice(n_samples, self.max_samples_for_fit, replace=False)
            X_fit = X[idx]
        else:
            X_fit = X

        self.pca_ = PCA(n_components=n_components, random_state=self.random_state)
        self.pca_.fit(X_fit)

        loadings = np.abs(self.pca_.components_)
        if self.weight_by_variance and hasattr(self.pca_, "explained_variance_ratio_"):
            weights = np.asarray(self.pca_.explained_variance_ratio_, dtype=np.float32).reshape(-1, 1)
            loadings = loadings * weights
        score = np.sum(loadings, axis=0)
        self.loading_scores_ = np.asarray(score, dtype=np.float32)
        self.selected_bands_ = np.argsort(-self.loading_scores_)[:n_select].astype(np.int64)
        self.selected_bands_sorted_ = np.sort(self.selected_bands_)
        self.n_features_in_ = int(n_bands)
        self.n_components_used_ = int(n_components)

        print(f"[PCA-LOADINGS] selected {len(self.selected_bands_)} / {n_bands} bands using {n_components} PCs:")
        print(self.selected_bands_.tolist())
        return self

    def transform(self, X):
        X = np.asarray(X, dtype=np.float32)
        if not hasattr(self, "selected_bands_"):
            raise RuntimeError("PCALoadingBandSelector is not fitted yet.")
        if X.ndim != 2:
            raise ValueError(f"PCALoadingBandSelector expects 2D X, got shape={X.shape}")
        if X.shape[1] != self.n_features_in_:
            raise ValueError(
                f"Band-count mismatch in PCALoadingBandSelector: got {X.shape[1]}, "
                f"expected {self.n_features_in_}."
            )
        return X[:, self.selected_bands_]

    def get_selected_bands(self):
        if not hasattr(self, "selected_bands_"):
            return None
        return self.selected_bands_.copy()

    def get_loading_scores(self):
        if not hasattr(self, "loading_scores_"):
            return None
        return self.loading_scores_.copy()



class SpectralAngleMapperClassifier(BaseEstimator, ClassifierMixin):
    """Simple Spectral Angle Mapper classifier.

    The reference spectrum for each class is the mean spectrum of that class in
    whatever feature space reaches this classifier. Therefore, when used inside
    a sklearn Pipeline, SavGol/SNV/MRMR/etc. are applied before class means are
    computed during fit() and before angles are computed during predict().

    Confidence source for the runtime threshold:
        predict_proba() = softmax(-angle / temperature)
    where smaller angle means better spectral match.
    """

    def __init__(self, temperature=0.05, eps=1e-12):
        self.temperature = float(temperature)
        self.eps = float(eps)

    def fit(self, X, y):
        X = np.asarray(X, dtype=np.float32)
        y = np.asarray(y)
        if X.ndim != 2:
            raise ValueError(f"SpectralAngleMapperClassifier expects 2D X, got shape={X.shape}")
        classes = np.unique(y)
        if classes.size == 0:
            raise ValueError("SpectralAngleMapperClassifier received no classes.")

        refs = []
        for cls in classes:
            mask = y == cls
            if not np.any(mask):
                continue
            refs.append(np.mean(X[mask], axis=0))

        self.classes_ = classes.astype(y.dtype, copy=False)
        self.references_ = np.asarray(refs, dtype=np.float32)
        self.n_features_in_ = int(X.shape[1])
        self.reference_norms_ = np.linalg.norm(self.references_, axis=1) + self.eps
        self.references_unit_ = self.references_ / self.reference_norms_[:, None]
        return self

    def _angles(self, X):
        X = np.asarray(X, dtype=np.float32)
        if X.ndim != 2:
            raise ValueError(f"SpectralAngleMapperClassifier expects 2D X, got shape={X.shape}")
        if not hasattr(self, "references_unit_"):
            raise RuntimeError("SpectralAngleMapperClassifier is not fitted yet.")
        if X.shape[1] != self.n_features_in_:
            raise ValueError(
                f"Band-count mismatch in SpectralAngleMapperClassifier: got {X.shape[1]}, "
                f"expected {self.n_features_in_}."
            )
        norms = np.linalg.norm(X, axis=1) + self.eps
        X_unit = X / norms[:, None]
        cos = X_unit @ self.references_unit_.T
        cos = np.clip(cos, -1.0, 1.0)
        return np.arccos(cos)

    def predict(self, X):
        angles = self._angles(X)
        best = np.argmin(angles, axis=1)
        return self.classes_[best]

    def decision_function(self, X):
        # Higher score is better for sklearn-style decision scores.
        return -self._angles(X)

    def predict_proba(self, X):
        angles = self._angles(X)
        temp = max(float(self.temperature), self.eps)
        scores = -angles / temp
        scores = scores - np.max(scores, axis=1, keepdims=True)
        exp_scores = np.exp(scores)
        denom = np.sum(exp_scores, axis=1, keepdims=True) + self.eps
        return exp_scores / denom

    def get_reference_spectra(self):
        if not hasattr(self, "references_"):
            return None
        return self.references_.copy()


class ConfidenceRejectingClassifier(BaseEstimator, ClassifierMixin):
    """Runtime-adjustable confidence rejection wrapper.

    Class convention:
        0 = Background / ignore
        1..N = material classes
        N+1 = Not classified / uncertain

    The wrapped estimator predicts only material classes. Low-confidence material
    predictions are remapped to reject_label. The threshold is intentionally a
    normal mutable attribute:

        pipeline.threshold = 0.85

    so SenSoRTC can adjust it live without retraining or reloading the model.

    Confidence source:
        - predict_proba(): max class probability
        - decision_function(): softmax-normalised decision scores
        - fallback: no rejection, because no confidence estimate exists
    """

    def __init__(self, estimator, threshold=0.70, reject_label=None):
        self.estimator = estimator
        self.threshold = self._normalise_threshold(threshold)
        self.reject_label = None if reject_label is None else int(reject_label)

    @staticmethod
    def _normalise_threshold(threshold):
        """Return a safe 0..1 confidence threshold.

        A threshold of 0.0 is the explicit OFF switch for the
        Not-classified confidence rejector.  This is important for live
        testing: with confidence set to 0, this wrapper must never create
        reject-label pixels by itself.
        """
        try:
            threshold = float(threshold)
        except Exception:
            return 0.0
        if not np.isfinite(threshold):
            return 0.0
        return float(np.clip(threshold, 0.0, 1.0))

    def fit(self, X, y):
        self.estimator.fit(X, y)
        self.classes_ = getattr(self.estimator, "classes_", np.unique(y))
        return self

    @staticmethod
    def _classes_for_score_columns(classes, n_columns):
        """Return one class label per score/proba column.

        Some sklearn estimators trained on a single material class expose
        classes_ with length 1, while their predict_proba/decision output can
        still have two columns.  Indexing classes_[1] then raises:
            index 1 is out of bounds for axis 0 with size 1

        For a one-material foreground classifier, both score columns still mean
        the same material label for our use case, so duplicate the sole class.
        """
        classes = np.asarray(classes)
        n_columns = int(max(1, n_columns))
        if classes.size == n_columns:
            return classes
        if classes.size == 1:
            return np.full((n_columns,), classes[0], dtype=classes.dtype)
        if classes.size == 0:
            return np.arange(n_columns, dtype=np.uint8)
        # Last-resort fallback for inconsistent estimator metadata.
        return np.resize(classes, n_columns)

    def _confidence_from_predict_proba(self, X):
        proba = self.estimator.predict_proba(X)
        proba = np.asarray(proba, dtype=np.float32)
        if proba.ndim == 1:
            proba = proba.reshape(-1, 1)

        classes = self._classes_for_score_columns(
            getattr(self.estimator, "classes_", []),
            proba.shape[1],
        )

        best_idx = np.argmax(proba, axis=1)
        pred = classes[best_idx]
        conf = proba[np.arange(proba.shape[0]), best_idx]
        return pred, conf

    def _confidence_from_decision_function(self, X):
        scores = self.estimator.decision_function(X)
        scores = np.asarray(scores, dtype=np.float32)

        raw_classes = getattr(self.estimator, "classes_", [])

        if scores.ndim == 1:
            # Binary classifier: sklearn returns signed distance for positive class.
            # Convert to two pseudo-probabilities with a logistic transform.
            classes = self._classes_for_score_columns(raw_classes, 2)

            p_pos = 1.0 / (1.0 + np.exp(-scores))
            p_neg = 1.0 - p_pos
            proba = np.vstack([p_neg, p_pos]).T
            best_idx = np.argmax(proba, axis=1)
            pred = classes[best_idx]
            conf = proba[np.arange(proba.shape[0]), best_idx]
            return pred, conf

        classes = self._classes_for_score_columns(raw_classes, scores.shape[1])

        # Softmax over decision scores gives an approximate confidence for
        # estimators without predict_proba, e.g. LinearSVC.
        shifted = scores - np.max(scores, axis=1, keepdims=True)
        exp_scores = np.exp(shifted)
        proba = exp_scores / (np.sum(exp_scores, axis=1, keepdims=True) + 1e-12)

        best_idx = np.argmax(proba, axis=1)
        pred = classes[best_idx]
        conf = proba[np.arange(proba.shape[0]), best_idx]
        return pred, conf

    def set_threshold(self, threshold):
        self.threshold = self._normalise_threshold(threshold)
        return self

    def get_threshold(self):
        return float(self.threshold)

    def set_reject_label(self, reject_label):
        self.reject_label = int(reject_label)
        return self

    def get_reject_label(self):
        if self.reject_label is None:
            classes = np.asarray(getattr(self.estimator, "classes_", []), dtype=np.int64)
            return int(classes.max()) + 1 if classes.size else 255
        return int(self.reject_label)

    def predict_with_confidence(self, X):
        try:
            if hasattr(self.estimator, "predict_proba"):
                pred, conf = self._confidence_from_predict_proba(X)
            elif hasattr(self.estimator, "decision_function"):
                pred, conf = self._confidence_from_decision_function(X)
            else:
                pred = self.estimator.predict(X)
                conf = np.ones((len(pred),), dtype=np.float32)
        except Exception as exc:
            # Confidence metadata can be inconsistent for one-class models.
            # Do not kill live acquisition; fall back to plain prediction and
            # disable rejection for this call by assigning full confidence.
            pred = self.estimator.predict(X)
            conf = np.ones((len(pred),), dtype=np.float32)

        pred = np.asarray(pred).astype(np.uint8, copy=False)
        conf = np.asarray(conf, dtype=np.float32)

        threshold = self._normalise_threshold(getattr(self, "threshold", 0.0))

        # threshold == 0.0 means confidence rejection is disabled.
        # Do not create any Not-classified pixels in that mode.
        if threshold <= 0.0:
            return pred.copy(), conf

        reject_label = self.get_reject_label()
        conf_for_reject = np.nan_to_num(
            conf,
            nan=-np.inf,
            posinf=1.0,
            neginf=-np.inf,
        )

        rejected = conf_for_reject < threshold
        pred = pred.copy()
        pred[rejected] = reject_label

        return pred, conf

    def predict(self, X):
        pred, _ = self.predict_with_confidence(X)
        return pred

    def predict_proba(self, X):
        if hasattr(self.estimator, "predict_proba"):
            return self.estimator.predict_proba(X)
        raise AttributeError("Wrapped estimator does not provide predict_proba().")

    def decision_function(self, X):
        if hasattr(self.estimator, "decision_function"):
            return self.estimator.decision_function(X)
        raise AttributeError("Wrapped estimator does not provide decision_function().")




class NIRPipelineBase:
    """Base class for SenSoRTC/EVK NIR sklearn classifiers."""

    COMMON_BAND_COUNTS = (220,212)

    def __init__(
        self,
        window_length=15,
        polyorder=2,
        deriv=1,
        random_state=42,
        confidence_threshold=0.70,
        reject_label=None,
    ):
        self.window_length = window_length
        self.polyorder = polyorder
        self.deriv = deriv
        self.random_state = random_state
        self.confidence_threshold = float(confidence_threshold)
        self.reject_label = None if reject_label is None else int(reject_label)

        self.X_train, self.X_test = None, None
        self.y_train, self.y_test = None, None
        self.class_names = None
        self.expected_bands = None

        self.pipe = self.build_pipeline()

    def build_pipeline(self):
        """Subclasses must return an sklearn Pipeline."""
        raise NotImplementedError

    def preprocessing_steps(self):
        """Reusable preprocessing for all NIR spectral classifiers."""
        return [
            ("savgol", SavGolTransformer(
                window_length=self.window_length,
                polyorder=self.polyorder,
                deriv=self.deriv,
            )),
            ("zscore", StandardScaler()),
        ]

    def preprocessing_metadata(self):
        """Describe the actual fitted sklearn Pipeline steps saved in the joblib.

        This is intentionally derived from self.pipe, because MRMR pipelines can
        now include SavGol/Z-score before band selection.  The metadata must
        describe the real pipeline, not a generic assumption.
        """
        try:
            steps = list(self.pipe.named_steps.keys())
        except Exception:
            steps = []
        return {
            "pipeline_steps": steps,
            "savgol_enabled": "savgol" in steps,
            "savgol_window_length": self.window_length if "savgol" in steps else None,
            "savgol_polyorder": self.polyorder if "savgol" in steps else None,
            "savgol_deriv": self.deriv if "savgol" in steps else None,
            "zscore_before_mrmr": "zscore_before_mrmr" in steps,
            "zscore_after_mrmr": "zscore_after_mrmr" in steps,
            "zscore": "zscore" in steps,
            "mrmr_enabled": "mrmr" in steps,
        }

    @staticmethod
    def _material_name_from_path(path):
        stem = Path(path).stem
        if stem.endswith("_timestamps"):
            raise ValueError(f"Do not pass timestamp files into training: {path}")
        return stem

    @staticmethod
    def _first_numeric_array_from_npz(npz, source_name):
        preferred = ("imnData", "data", "spectra", "X", "arr_0")
        for key in preferred:
            if key in npz.files:
                return npz[key]
        for key in npz.files:
            if not str(key).lower().endswith("timestamps"):
                return npz[key]
        raise ValueError(f"No spectral array found in NPZ file: {source_name}")

    @staticmethod
    def _numeric_excel_matrix(path, sheet_name=None):
        """Read an Excel sheet and return a dense numeric matrix.

        Rows with fewer than two numeric values are ignored. This allows files
        with one text/header column or header rows. Non-numeric cells inside a
        numeric row become NaN and columns that are completely NaN are removed.
        """
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise ImportError("Excel loading requires openpyxl") from exc

        wb = load_workbook(path, data_only=True, read_only=True)
        if sheet_name is None:
            ws = wb[wb.sheetnames[0]]
        else:
            ws = wb[sheet_name]

        rows = []
        for row in ws.iter_rows(values_only=True):
            numeric = []
            n_numeric = 0
            for value in row:
                if isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool):
                    numeric.append(float(value))
                    n_numeric += 1
                else:
                    numeric.append(np.nan)
            if n_numeric >= 2:
                rows.append(numeric)

        wb.close()

        if not rows:
            raise ValueError(f"No numeric spectra found in Excel file: {path}")

        max_len = max(len(r) for r in rows)
        arr = np.full((len(rows), max_len), np.nan, dtype=np.float32)
        for i, row in enumerate(rows):
            arr[i, :len(row)] = row

        # Drop columns that are entirely non-numeric.
        valid_cols = ~np.all(np.isnan(arr), axis=0)
        arr = arr[:, valid_cols]

        # Drop rows that still contain NaN after removing non-numeric columns.
        # This is intentionally strict: each remaining row must be one complete spectrum.
        valid_rows = ~np.any(np.isnan(arr), axis=1)
        arr = arr[valid_rows]

        if arr.size == 0:
            raise ValueError(f"No complete numeric spectra found in Excel file: {path}")

        return arr.astype(np.float32, copy=False)

    def _normalise_2d_spectra_matrix(self, arr, source_name="array"):
        """Convert a 2-D numeric matrix to X=(n_spectra, bands).

        Preferred convention: rows are spectra and columns are bands.
        If the matrix appears transposed, e.g. (220, n_spectra), it is transposed.
        """
        arr = np.asarray(arr, dtype=np.float32)
        if arr.ndim != 2:
            raise ValueError(f"Expected 2D matrix for {source_name}, got shape={arr.shape}")

        if self.expected_bands is not None:
            if arr.shape[1] == self.expected_bands:
                return arr
            if arr.shape[0] == self.expected_bands:
                return arr.T
            raise ValueError(
                f"Band-count mismatch for {source_name}: shape={arr.shape}, "
                f"expected {self.expected_bands} bands."
            )

        # First file: infer orientation. Prefer known NIR band counts if present.
        for bands in self.COMMON_BAND_COUNTS:
            if arr.shape[1] == bands:
                return arr
            if arr.shape[0] == bands:
                return arr.T

        # Generic fallback: assume rows are spectra, columns are bands.
        return arr

    def _array_to_spectra_legacy_or_evk(self, arr, source_name="array"):
        """Convert spectral arrays to X=(n_spectra, bands).

        EVK SQALAR / new SenSoRTC:
            (width, bands, lines), e.g. (312, 220, nLines)

        Old SenSoRTC before transpose change:
            (lines, width, bands), e.g. (nLines, 312, 220)

        2-D matrices are handled by _normalise_2d_spectra_matrix().
        """
        arr = np.asarray(arr)

        if arr.ndim == 3:
            shape = arr.shape

            if self.expected_bands is not None:
                if shape[1] == self.expected_bands:  # EVK/new SenSoRTC
                    bands = shape[1]
                    return arr.transpose(0, 2, 1).reshape(-1, bands)
                if shape[2] == self.expected_bands:  # old SenSoRTC
                    bands = shape[2]
                    return arr.reshape(-1, bands)
                if shape[0] == self.expected_bands:
                    bands = shape[0]
                    moved = np.moveaxis(arr, 0, -1)
                    return moved.reshape(-1, bands)
                raise ValueError(
                    f"Band-count mismatch for {source_name}: shape={shape}, "
                    f"expected {self.expected_bands} bands."
                )

            # EVK/new SenSoRTC: (width, bands, lines)
            for bands in self.COMMON_BAND_COUNTS:
                if shape[1] == bands:
                    return arr.transpose(0, 2, 1).reshape(-1, bands)

            # Old SenSoRTC: (lines, width, bands)
            for bands in self.COMMON_BAND_COUNTS:
                if shape[2] == bands:
                    return arr.reshape(-1, bands)

            # Fallback: choose a plausible spectral axis. Prefer middle axis.
            candidate_axes = [i for i, s in enumerate(shape) if 16 <= s <= 4096]
            if not candidate_axes:
                raise ValueError(f"Cannot infer spectral axis for {source_name}: {shape}")

            spectral_axis = 1 if 1 in candidate_axes else candidate_axes[-1]
            bands = shape[spectral_axis]
            moved = np.moveaxis(arr, spectral_axis, -1)
            return moved.reshape(-1, bands)

        if arr.ndim == 2:
            return self._normalise_2d_spectra_matrix(arr, source_name=source_name)

        raise ValueError(f"Unsupported ndim for {source_name}: {arr.ndim}, shape={arr.shape}")

    def _filter_background(self, X, min_mean):
        X = np.asarray(X, dtype=np.float32)
        if min_mean is None:
            return X
        keep = X.mean(axis=1) >= float(min_mean)
        X = X[keep]
        if X.size == 0:
            raise ValueError(
                f"No spectra left after min_mean={min_mean}. "
                "Lower min_mean or check whether this file contains foreground material."
            )
        return X

    def _finish_loaded_X_y(self, X, label, min_mean, source_name):        

        if self.expected_bands is None:
            self.expected_bands = int(X.shape[1])
        elif X.shape[1] != self.expected_bands:
            raise ValueError(
                f"Band-count mismatch after loading {source_name}: "
                f"got {X.shape[1]}, expected {self.expected_bands}."
            )

        y = np.full(X.shape[0], label, dtype=np.uint8)
        return X, y

    def load_imnData_mat(self, path, label, min_mean=800):
        path = Path(path)
        mat = loadmat(path)

        if "imnData" in mat:
            arr = mat["imnData"]
        else:
            keys = [k for k in mat.keys() if not k.startswith("__")]
            if not keys:
                raise ValueError(f"No data arrays found in MAT file: {path}")
            arr = mat[keys[0]]

        X = self._array_to_spectra_legacy_or_evk(arr, source_name=str(path))
        X = self._filter_background(X, min_mean=min_mean)
        return self._finish_loaded_X_y(X, label, min_mean, source_name=str(path))

    def load_npy_spectral(self, path, label, min_mean=800):
        path = Path(path)
        if path.stem.endswith("_timestamps"):
            raise ValueError(f"This is a timestamp file, not spectral data: {path}")

        arr = np.load(path)
        X = self._array_to_spectra_legacy_or_evk(arr, source_name=str(path))
        X = self._filter_background(X, min_mean=min_mean)
        return self._finish_loaded_X_y(X, label, min_mean, source_name=str(path))

    def load_npz_spectral(self, path, label, min_mean=800):
        path = Path(path)
        with np.load(path) as npz:
            arr = self._first_numeric_array_from_npz(npz, source_name=str(path))
        X = self._array_to_spectra_legacy_or_evk(arr, source_name=str(path))
        X = self._filter_background(X, min_mean=min_mean)
        return self._finish_loaded_X_y(X, label, min_mean, source_name=str(path))

    def load_excel_spectral(self, path, label, min_mean=800, sheet_name=None):
        path = Path(path)
        arr = self._numeric_excel_matrix(path, sheet_name=sheet_name)
        X = self._normalise_2d_spectra_matrix(arr, source_name=str(path))
        return self._finish_loaded_X_y(X, label, min_mean, source_name=str(path))

    def load_spectral_file(self, path, label, min_mean=800):
        path = Path(path)
        suffix = path.suffix.lower()

        if suffix == ".mat":
            return self.load_imnData_mat(path, label, min_mean=min_mean)
        if suffix == ".npy":
            return self.load_npy_spectral(path, label, min_mean=min_mean)
        if suffix == ".npz":
            return self.load_npz_spectral(path, label, min_mean=min_mean)
        if suffix in (".xlsx", ".xlsm"):
            return self.load_excel_spectral(path, label, min_mean=min_mean)
        if suffix == ".xls":
            raise ValueError(".xls is not supported by openpyxl. Save as .xlsx or .xlsm.")

        raise ValueError(f"Unsupported file type: {path}")

    def load_data(self, paths, min_mean=800, test_size=0.2, random_state=None):
        if random_state is None:
            random_state = self.random_state

        self.expected_bands = None
        X_parts = []
        y_parts = []

        for class_idx, path in enumerate(paths, start=1):
            X_data, y_data = self.load_spectral_file(path, class_idx, min_mean=min_mean)
            print(f"Loaded {Path(path).name}: X={X_data.shape}, label={class_idx}")
            X_parts.append(X_data)
            y_parts.append(y_data)

        X = np.vstack(X_parts)
        y = np.concatenate(y_parts)

        self.X_train, self.X_test, self.y_train, self.y_test = train_test_split(
            X,
            y,
            test_size=test_size,
            stratify=y,
            random_state=random_state,
        )

    def train(self, paths, name=None, min_mean=800, test_size=0.2, confidence_threshold=None):
        if name is None:
            name = f"{self.__class__.__name__}.joblib"

        self.load_data(paths, min_mean=min_mean, test_size=test_size)

        if confidence_threshold is not None:
            self.confidence_threshold = float(confidence_threshold)

        self.pipe.fit(self.X_train, self.y_train)

        pred = self.pipe.predict(self.X_test)
        print(confusion_matrix(self.y_test, pred))

        material_names = [self._material_name_from_path(p) for p in paths]
        print(classification_report(
            self.y_test,
            pred,
            labels=list(range(1, len(paths) + 1)),
            target_names=material_names,
        ))

        not_classified_label = int(self.reject_label) if self.reject_label is not None else len(paths) + 1

        wrapped_pipe = ConfidenceRejectingClassifier(
            self.pipe,
            threshold=self.confidence_threshold,
            reject_label=not_classified_label,
        )

        selected_bands = None
        try:
            mrmr = self.pipe.named_steps.get("mrmr")
            if hasattr(mrmr, "get_selected_bands"):
                selected = mrmr.get_selected_bands()
                if selected is not None:
                    selected_bands = [int(v) for v in selected]
        except Exception:
            selected_bands = None

        bundle = {
            "pipeline": wrapped_pipe,
            "class_names": ["Background"] + material_names + ["Not classified"],
            "class_labels": list(range(len(paths) + 2)),
            "kind": self.__class__.__name__,
            "format": "sklearn_pipeline_bundle_v4_savgol_zscore_mrmr",
            "n_bands": self.expected_bands,
            "selected_bands": selected_bands,
            "n_selected_bands": None if selected_bands is None else len(selected_bands),

            # Runtime-adjustable threshold metadata.
            # SenSoRTC should use this as the initial slider value, then write
            # live values to bundle["pipeline"].threshold.
            "recommended_confidence_threshold": self.confidence_threshold,
            "confidence_threshold": self.confidence_threshold,
            "confidence_threshold_runtime_adjustable": True,
            "confidence_threshold_attribute": "threshold",

            "reject_label": not_classified_label,
            "background_label": 0,
            "not_classified_label": not_classified_label,
            "preprocessing": self.preprocessing_metadata(),
        }

        joblib.dump(bundle, name)
        print(f"Saved: {name}")
        return bundle


class Shallow_NN(NIRPipelineBase):
    def __init__(
        self,
        window_length=15,
        polyorder=2,
        deriv=1,
        hidden_layer_sizes=(64, 32),
        random_state=42,
        confidence_threshold=0.70,
        reject_label=None,
    ):
        self.hidden_layer_sizes = hidden_layer_sizes
        super().__init__(
            window_length=window_length,
            polyorder=polyorder,
            deriv=deriv,
            random_state=random_state,
            confidence_threshold=confidence_threshold,
            reject_label=reject_label,
        )

    def build_pipeline(self):
        return Pipeline(
            self.preprocessing_steps() + [
                ("mlp", MLPClassifier(
                    hidden_layer_sizes=self.hidden_layer_sizes,
                    activation="relu",
                    alpha=1e-4,
                    learning_rate_init=1e-3,
                    max_iter=500,
                    early_stopping=True,
                    random_state=self.random_state,
                )),
            ]
        )


class SVM_RBF(NIRPipelineBase):
    def __init__(
        self,
        window_length=15,
        polyorder=2,
        deriv=1,
        C=10.0,
        gamma="scale",
        class_weight="balanced",
        random_state=42,
        confidence_threshold=0.70,
        reject_label=None,
    ):
        self.C = C
        self.gamma = gamma
        self.class_weight = class_weight
        super().__init__(
            window_length=window_length,
            polyorder=polyorder,
            deriv=deriv,
            random_state=random_state,
            confidence_threshold=confidence_threshold,
            reject_label=reject_label,
        )

    def build_pipeline(self):
        return Pipeline(
            self.preprocessing_steps() + [
                ("svm", SVC(
                    kernel="rbf",
                    C=self.C,
                    gamma=self.gamma,
                    class_weight=self.class_weight,
                    probability=True,
                )),
            ]
        )


class SVM_Linear(NIRPipelineBase):
    def __init__(
        self,
        window_length=15,
        polyorder=2,
        deriv=1,
        C=1.0,
        class_weight="balanced",
        random_state=42,
        confidence_threshold=0.70,
        reject_label=None,
    ):
        self.C = C
        self.class_weight = class_weight
        super().__init__(
            window_length=window_length,
            polyorder=polyorder,
            deriv=deriv,
            random_state=random_state,
            confidence_threshold=confidence_threshold,
            reject_label=reject_label,
        )

    def build_pipeline(self):
        return Pipeline(
            self.preprocessing_steps() + [
                ("linear_svm", LinearSVC(
                    C=self.C,
                    class_weight=self.class_weight,
                    random_state=self.random_state,
                    max_iter=10000,
                )),
            ]
        )


class SVM_Linear_MRMR(NIRPipelineBase):
    """Fast NIR classifier: SavGol/Z-score + mRMR band selection + LinearSVC.

    The live camera still provides full raw spectra.  The saved sklearn Pipeline
    applies the same preprocessing during inference as during training:
        raw spectrum -> SavGol -> Z-score -> MRMR -> classifier

    The selected band indices refer to the spectral columns after SavGol/Z-score.
    SavGol and z-score do not change the number/order of spectral columns, so
    the indices still correspond to original band positions.
    """

    def __init__(
        self,
        n_bands_select=30,
        redundancy_weight=1.0,
        C=1.0,
        class_weight="balanced",
        window_length=15,
        polyorder=2,
        deriv=1,
        use_savgol_before_mrmr=True,
        use_zscore_before_mrmr=True,
        use_zscore_after_mrmr=False,
        random_state=42,
        confidence_threshold=0.70,
        reject_label=None,
    ):
        self.n_bands_select = int(n_bands_select)
        self.redundancy_weight = float(redundancy_weight)
        self.C = C
        self.class_weight = class_weight
        self.use_savgol_before_mrmr = bool(use_savgol_before_mrmr)
        self.use_zscore_before_mrmr = bool(use_zscore_before_mrmr)
        self.use_zscore_after_mrmr = bool(use_zscore_after_mrmr)
        super().__init__(
            window_length=window_length,
            polyorder=polyorder,
            deriv=deriv,
            random_state=random_state,
            confidence_threshold=confidence_threshold,
            reject_label=reject_label,
        )

    def _mrmr_steps(self):
        steps = []
        if self.use_savgol_before_mrmr:
            steps.append(("savgol", SavGolTransformer(
                window_length=self.window_length,
                polyorder=self.polyorder,
                deriv=self.deriv,
            )))
        if self.use_zscore_before_mrmr:
            steps.append(("zscore_before_mrmr", StandardScaler()))
        steps.append(("mrmr", MRMRBandSelector(
            n_features=self.n_bands_select,
            redundancy_weight=self.redundancy_weight,
            random_state=self.random_state,
        )))
        if self.use_zscore_after_mrmr:
            steps.append(("zscore_after_mrmr", StandardScaler()))
        return steps

    def build_pipeline(self):
        return Pipeline(
            self._mrmr_steps() + [
                ("linear_svm", LinearSVC(
                    C=self.C,
                    class_weight=self.class_weight,
                    random_state=self.random_state,
                    max_iter=10000,
                )),
            ]
        )

class Shallow_NN_MRMR(NIRPipelineBase):
    """SavGol/Z-score + mRMR band selection + small MLP.

    Usually slower than SVM_Linear_MRMR but still much lighter than using all
    spectral bands.  The saved Pipeline contains all preprocessing, so the NIR
    runtime can call pipeline.predict(raw_spectra) directly.
    """

    def __init__(
        self,
        n_bands_select=30,
        redundancy_weight=1.0,
        hidden_layer_sizes=(24,),
        window_length=15,
        polyorder=2,
        deriv=1,
        use_savgol_before_mrmr=True,
        use_zscore_before_mrmr=True,
        use_zscore_after_mrmr=False,
        random_state=42,
        confidence_threshold=0.70,
        reject_label=None,
    ):
        self.n_bands_select = int(n_bands_select)
        self.redundancy_weight = float(redundancy_weight)
        self.hidden_layer_sizes = hidden_layer_sizes
        self.use_savgol_before_mrmr = bool(use_savgol_before_mrmr)
        self.use_zscore_before_mrmr = bool(use_zscore_before_mrmr)
        self.use_zscore_after_mrmr = bool(use_zscore_after_mrmr)
        super().__init__(
            window_length=window_length,
            polyorder=polyorder,
            deriv=deriv,
            random_state=random_state,
            confidence_threshold=confidence_threshold,
            reject_label=reject_label,
        )

    def _mrmr_steps(self):
        steps = []
        if self.use_savgol_before_mrmr:
            steps.append(("savgol", SavGolTransformer(
                window_length=self.window_length,
                polyorder=self.polyorder,
                deriv=self.deriv,
            )))
        if self.use_zscore_before_mrmr:
            steps.append(("zscore_before_mrmr", StandardScaler()))
        steps.append(("mrmr", MRMRBandSelector(
            n_features=self.n_bands_select,
            redundancy_weight=self.redundancy_weight,
            random_state=self.random_state,
        )))
        if self.use_zscore_after_mrmr:
            steps.append(("zscore_after_mrmr", StandardScaler()))
        return steps

    def build_pipeline(self):
        return Pipeline(
            self._mrmr_steps() + [
                ("mlp", MLPClassifier(
                    hidden_layer_sizes=self.hidden_layer_sizes,
                    activation="relu",
                    alpha=1e-4,
                    learning_rate_init=1e-3,
                    max_iter=500,
                    early_stopping=True,
                    random_state=self.random_state,
                )),
            ]
        )
